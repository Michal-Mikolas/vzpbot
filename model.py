import os
import re
import openpyxl
from datetime import datetime
from dateutil.parser import parse


#######
#       #    #  ####  ###### #####  ##### #  ####  #    #  ####
#        #  #  #    # #      #    #   #   # #    # ##   # #
#####     ##   #      #####  #    #   #   # #    # # #  #  ####
#         ##   #      #      #####    #   # #    # #  # #      #
#        #  #  #    # #      #        #   # #    # #   ## #    #
####### #    #  ####  ###### #        #   #  ####  #    #  ####

class TooBigFile(Exception):
	pass

class NotSupportedExtension(Exception):
	pass


#     #
##   ##  ####  #####  ###### #
# # # # #    # #    # #      #
#  #  # #    # #    # #####  #
#     # #    # #    # #      #
#     # #    # #    # #      #
#     #  ####  #####  ###### ######

class Model(object):

	def __init__(self, file):
		self.wb = None
		self.ws = None
		self.r = 1  # starting from row 2, first is header
		self.max_row = 0

		# Only XLSX is supported
		ext = file.split(".")[-1].lower()
		if ext not in ("xlsx"):
			raise NotSupportedExtension('The file has to be .xlsx')

		# Size check
		size = os.path.getsize(file)
		if size > 8000000:
			raise TooBigFile(file)

		# Opene excel
		self.wb = openpyxl.load_workbook(file, data_only=True)
		self.ws = self.wb.active
		self.max_row = self.ws.max_row

	def next(self):
		self.r += 1
		if self.r > self.max_row:
			return None

		return {
			'r': self.r,
			'surname': self._fix_str(self.ws.cell(self.r, 1).value),
			'name': self._fix_str(self.ws.cell(self.r, 2).value),
			'birthdate': self._parse_date(self.ws.cell(self.r, 3).value),
			'insurance': self._fix_str(self.ws.cell(self.r, 8).value),
		}

	def persist(self, entry):
		self.ws.cell(entry['r'], 8).value = entry['insurance']

	def save(self, file):
		self.wb.save(file)


	###
	 #  #    # ##### ###### #####  #    #   ##   #
	 #  ##   #   #   #      #    # ##   #  #  #  #
	 #  # #  #   #   #####  #    # # #  # #    # #
	 #  #  # #   #   #      #####  #  # # ###### #
	 #  #   ##   #   #      #   #  #   ## #    # #
	### #    #   #   ###### #    # #    # #    # ######

	def _fix_str(self, value, strip = True):
		value = str(value).strip()

		if strip:
			value = value.strip()

		if re.search(r'^\d+\.0$', value):  # revert excel 'decimal to float' conversion
			value = re.sub(r'\.0$', '', value)

		return value

	def _parse_date(self, date):
		if date and type(date).__name__ == 'str':
			# dd.mm.yyyy
			d = re.search(r'^(\d+)\.\s*(\d+)\.\s?(\d+)$', date)
			if d:
				year = int(d[3])
				month = int(d[2])
				day = int(d[1])

				if (day > 31) or (month > 12):
					raise ValueError()

				return datetime(year, month, day)

			# mm/dd/yyyy
			d = re.search(r'^(\d+)/(\d+)/(\d+)$', date)
			if d:
				year = int(d[3])
				month = int(d[1])
				day = int(d[2])

				if (day > 31) or (month > 12):
					raise ValueError()

				return datetime(year, month, day)

			# yyyy-mm-dd
			d = re.search(r'^(\d+)-(\d+)-(\d+)$', date)
			if d:
				year = int(d[1])
				month = int(d[2])
				day = int(d[3])

				if (day > 31) or (month > 12):
					raise ValueError()

				return datetime(year, month, day)

			return parse(date)

		return date
